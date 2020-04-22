import openpyxl

WB= openpyxl.load_workbook("C:\\Desktop\\TAX\\Python Pytest Training\\PracticeWB.xlsx")
Sheet= WB.active
Cell= Sheet.cell(1,4)
Cell_Value = Cell.value

Rows= Sheet.max_row
print(Sheet.max_row)

Columns= Sheet.max_column
print(Sheet.max_column)
print(Cell_Value)

Dict = {}

for i in range(2,Rows+1):
    for j in range(2, Columns+1):
        Dict[Sheet.cell(row=1, column=j).value] = Sheet.cell(row=i, column=j).value

    print(Dict)




